import streamlit as st
import pandas as pd
import tempfile
import os
from openpyxl import load_workbook
from openpyxl.styles import Font

# Mapping dictionary
police_station_mapping = {
    "11188003": "ભીલોડા", "11188010": "શામળાજી", "11188004": "ધનસુરા",
    "11188002": "બાયડ", "11188001": "આબલીયારા", "11188009": "મોડાસા_ટાઉન",
    "11188008": "મોડાસા_રૂરલ", "11188007": "મેધરજ", "11188006": "માલપુર",
    "11188005": "ઇસરી", "11188011": "સાથંબા", "11188012": "મહિલા_પોલીસ_સ્ટેશન",
    "11188013": "ટીંટોઇ", "11188014": "સાયબર ક્રાઇમ પોલીસ સ્ટેશન",
}

st.title("📊 FIR & SID Excel Processor")

mode = st.selectbox("Select Processing Mode", ["Fir Link SID", "Fir ma use karel SID"])

# Upload SID folder (as multiple files)
sid_files = st.file_uploader("Upload all SID .xls files", accept_multiple_files=True, type=["xls"])

# Upload FIR Case file
fir_file = st.file_uploader("Upload FIR Case.xls file", type=["xls"])

if st.button("Generate Report") and sid_files and fir_file:
    with tempfile.TemporaryDirectory() as tmpdir:
        # Save SID files temporarily
        sid_paths = []
        for sid_file in sid_files:
            sid_path = os.path.join(tmpdir, sid_file.name)
            with open(sid_path, "wb") as f:
                f.write(sid_file.read())
            sid_paths.append(sid_path)

        # Save FIR file
        fir_path = os.path.join(tmpdir, "case.xls")
        with open(fir_path, "wb") as f:
            f.write(fir_file.read())

        # Load data
        sid_df_list = [pd.read_excel(p, engine='xlrd', header=None) for p in sid_paths]
        merged_sid_df = pd.concat(sid_df_list, ignore_index=True)
        df2 = pd.read_excel(fir_path, engine='xlrd', header=None)

        police_station_name = df2.iloc[4, 1]
        date_column = df2.iloc[4:, 2].dropna()
        start_date = pd.to_datetime(date_column.iloc[0], dayfirst=True).strftime("%d/%m/%Y")
        end_date = pd.to_datetime(date_column.iloc[-1], dayfirst=True).strftime("%d/%m/%Y")

        # Mode-specific logic
        case_number_1 = merged_sid_df.iloc[3:, 2].reset_index(drop=True)
        case_number_2 = merged_sid_df.iloc[3:, 10].reset_index(drop=True)
        fir_number = df2.iloc[4:, 1].reset_index(drop=True)

        if mode == "Fir Link SID":
            all_case_numbers = pd.concat([case_number_1, case_number_2]).dropna().unique()
            final_output = fir_number.apply(lambda x: x if x in all_case_numbers else None)
        else:  # "Fir ma use karel SID"
            all_fir_numbers = fir_number.dropna().unique()
            combined_sids = pd.concat([case_number_1, case_number_2]).dropna().unique()
            final_output = pd.Series([
                sid if sid in all_fir_numbers else None for sid in combined_sids
            ])

        output_df = pd.DataFrame({
            "Case_Number_1": case_number_1,
            "Case Number 2": case_number_2,
            "FIR Number": fir_number
        })

        output_df["Final Output"] = output_df["FIR Number"].apply(
            lambda x: x if x in all_case_numbers else None
        )

        output_df["Pending SID"] = output_df.apply(
            lambda row: row["FIR Number"] if pd.isna(row["Final Output"]) else None, axis=1
        )

        output_df["FIR Prefix"] = output_df["FIR Number"].astype(str).str[:8]
        output_df["Mapped Police Station"] = output_df["FIR Prefix"].map(police_station_mapping)

        io_map = dict(zip(df2.iloc[4:, 1], df2.iloc[4:, 6]))
        sheet2_data = []
        last_prefix = None
        output_df_sorted = output_df.sort_values(by=["FIR Prefix", "FIR Number"])
        for _, row in output_df_sorted.iterrows():
            fir_prefix = row["FIR Prefix"]
            station = row["Mapped Police Station"]
            fir_num = row["FIR Number"]
            final_out = row["Final Output"]
            pending = row["Pending SID"]
            pending_link = pending if pd.notna(pending) else None
            io_name = io_map.get(pending_link, "") if pending_link else ""
            sheet2_data.append([fir_prefix if fir_prefix != last_prefix else '', 
                                station if fir_prefix != last_prefix else '', 
                                fir_num, final_out, pending, pending_link, io_name])
            last_prefix = fir_prefix
        sheet2_df = pd.DataFrame(sheet2_data, columns=[
            "FIR Prefix", "Mapped Police Station", "FIR Number", "Final Output",
            "Pending SID", "Pending Fir Link", "IO Name"
        ])

        dashboard_data = []
        for station in output_df["Mapped Police Station"].dropna().unique():
            group = output_df[output_df["Mapped Police Station"] == station]
            fir_count = group["FIR Number"].count()
            final_count = group["Final Output"].count()
            pending_count = group["Pending SID"].count()
            percentage = round((final_count / fir_count) * 100, 2) if fir_count else 0
            dashboard_data.append([station, fir_count, final_count, pending_count, percentage])
        dashboard_df = pd.DataFrame(
            dashboard_data,
            columns=["પો.સ્ટેનુ નામ", "એફ.આઇ.આર સંખ્યા", "SID સંખ્યા", "SID બાકી સંખ્યા", "ટકાવારી"]
        )
        dashboard_df = dashboard_df.sort_values(by="ટકાવારી", ascending=False).reset_index(drop=True)
        dashboard_df.insert(0, "ક્રમ સં.", range(1, len(dashboard_df) + 1))
        title_row = pd.DataFrame([[
            f"E-Sakshya SID  Dt.{start_date} To Dt.{end_date}", None, None, None, None, None
        ]], columns=dashboard_df.columns)
        header_row = pd.DataFrame([dashboard_df.columns.tolist()], columns=dashboard_df.columns)
        total_row = pd.DataFrame([[
            "", "કુલ",
            dashboard_df["એફ.આઇ.આર સંખ્યા"].sum(),
            dashboard_df["SID સંખ્યા"].sum(),
            dashboard_df["SID બાકી સંખ્યા"].sum(),
            round((dashboard_df["SID સંખ્યા"].sum() / dashboard_df["એફ.આઇ.આર સંખ્યા"].sum()) * 100, 2)
        ]], columns=dashboard_df.columns)
        sheet3_df = pd.concat([title_row, header_row, dashboard_df, total_row], ignore_index=True)

        # Sheet4 Summary
        sheet4_df = pd.DataFrame({
            "વિશ્લેષણ મોડ": [mode],
            "કુલ FIR": [dashboard_df["એફ.આઇ.આર સંખ્યા"].sum()],
            "કુલ SID": [dashboard_df["SID સંખ્યા"].sum()],
            "બાકી SID": [dashboard_df["SID બાકી સંખ્યા"].sum()],
            "ટકાવારી": [round((dashboard_df["SID સંખ્યા"].sum() / dashboard_df["એફ.આઇ.આર સંખ્યા"].sum()) * 100, 2)]
        })

        # Save Excel
        output_path = os.path.join(tmpdir, "output.xlsx")
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            output_df.to_excel(writer, index=False, sheet_name="Sheet1")
            sheet2_df.to_excel(writer, index=False, sheet_name="Sheet2")
            sheet3_df.to_excel(writer, index=False, header=False, sheet_name="Sheet3")
            sheet4_df.to_excel(writer, index=False, sheet_name="Sheet4")

        # Format Sheet3 Bold
        wb = load_workbook(output_path)
        ws3 = wb["Sheet3"]
        bold_font = Font(bold=True)
        for cell in ws3[2]: cell.font = bold_font
        for cell in ws3[ws3.max_row]: cell.font = bold_font
        wb.save(output_path)

        # Download
        with open(output_path, "rb") as f:
            st.download_button("📥 Download Output Excel", f, file_name="Megh.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
